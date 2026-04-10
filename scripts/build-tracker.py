"""
build-tracker.py: Build the Midwife-Tracker.xlsx outreach workbook.

Reads outreach/calm-enriched.json (produced by scripts/scrape-calm.mjs) and
produces outreach/Midwife-Tracker.xlsx with 5 sheets:

  1. Targets          SoCal CALM midwives with contact info, priority scored
  2. Outreach Log     Row per send, tracks variant, opens, replies, outcomes
  3. Follow-ups       Formula-driven view showing days since last contact
  4. Message Variants 4 A/B/C/D email templates with merge fields
  5. Summary          Counts, variant performance, funnel

Run:
  pip install openpyxl --break-system-packages
  python3 scripts/build-tracker.py

Outputs outreach/Midwife-Tracker.xlsx (gitignored, contains PII).
"""

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
# Prefer the combined pool (CALM + Supabase + future sources) if it exists,
# fall back to the raw CALM roster for backwards compatibility.
COMBINED_INPUT = ROOT / "outreach" / "combined-targets.json"
CALM_INPUT = ROOT / "outreach" / "calm-enriched.json"
OUTPUT = ROOT / "outreach" / "Midwife-Tracker.xlsx"

# Design tokens lifted from the Homebirth Soft & Illustrated palette,
# rendered in Excel-safe hex (no alpha).
LAVENDER_DARK = "6B4C7C"
LAVENDER = "8B5FA0"
LAVENDER_LIGHT = "D4B8E0"
LAVENDER_LIGHTER = "F5E6F9"
PEACH = "F0CFC0"
GREEN = "B8D4A8"
PINK = "E8A0B8"
TEXT_DARK = "3A3535"
TEXT_MUTED = "7A6E6E"
BORDER_SOFT = "F0E4F5"

FONT_NAME = "Arial"

thin = Side(border_style="thin", color=BORDER_SOFT)
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)


def load_targets():
    # Prefer combined-targets.json when it exists. It already carries source
    # and tier and is pre-normalized by dedupe-calm-vs-supabase.mjs.
    if COMBINED_INPUT.exists():
        rows = json.loads(COMBINED_INPUT.read_text(encoding="utf-8"))
        # combined-targets.json rows already have the right shape. Backfill
        # defaults for any fields the dedupe script did not populate.
        for r in rows:
            r.setdefault("source", "Unknown")
            r.setdefault("tier", 3)
            r.setdefault("notes", "")
            r.setdefault("profile_url", "")
            r.setdefault("practice_details", "")
            r.setdefault("counties_all", r.get("primary_county") or "")
    else:
        # Legacy path: build from CALM only
        data = json.loads(CALM_INPUT.read_text(encoding="utf-8"))
        socal = [m for m in data if m.get("is_socal")]
        rows = []
        for m in socal:
            f = m.get("fields") or {}
            first = f.get("First name") or (m["name"].split(" ", 1)[0] if m.get("name") else "")
            last = f.get("Last name") or (m["name"].split(" ", 1)[1] if m.get("name") and " " in m["name"] else "")
            organization = f.get("Organization") or m.get("practice") or ""
            email = f.get("e-Mail") or m.get("email") or ""
            phone = f.get("Phone") or m.get("phone") or ""
            license_type = f.get("California Midwifery License Type") or ""
            license_no = f.get("License Number") or ""
            city = f.get("City") or ""
            counties_raw = f.get("Counties Served") or m.get("counties_served") or ""
            medi_cal = f.get("Medi-Cal") or m.get("medi_cal") or ""
            web = f.get("Web Address") or m.get("website") or ""
            practice_details = f.get("Practice Details") or ""
            counties_list = m.get("county_list") or []
            primary_county = counties_list[0] if counties_list else ""

            tier = 1 if email else (2 if web else 3)
            rows.append(
                {
                    "source": "CALM",
                    "tier": tier,
                    "first": first,
                    "last": last,
                    "full_name": m.get("name") or f"{first} {last}".strip(),
                    "organization": organization,
                    "license_type": license_type,
                    "license_no": license_no,
                    "city": city,
                    "primary_county": primary_county,
                    "counties_all": counties_raw,
                    "email": email,
                    "phone": phone,
                    "website": web,
                    "medi_cal": medi_cal,
                    "practice_details": practice_details,
                    "profile_url": m.get("profile_url") or "",
                    "notes": "",
                }
            )

    # Priority score: Tier 1 (has email) always beats Tier 2/3. Within tier,
    # LMs, orgs with websites, private-pay practices, and the big three
    # counties get bonus weight. Confirmed by two sources (CALM + Supabase)
    # gets an extra bump.
    def score(r):
        s = 0
        tier = int(r.get("tier") or 3)
        if tier == 1:
            s += 8
        elif tier == 2:
            s += 3
        if str(r.get("license_type") or "").upper().startswith("LM"):
            s += 2
        if r.get("organization"):
            s += 1
        if r.get("website"):
            s += 1
        if (r.get("medi_cal") or "").strip().lower() == "no":
            s += 1
        if r.get("primary_county") in ("Los Angeles", "Orange", "San Diego"):
            s += 1
        if r.get("source") == "CALM + Supabase":
            s += 1
        return s

    rows.sort(key=lambda r: (-score(r), (r.get("last") or "").lower(), (r.get("first") or "").lower()))
    for i, r in enumerate(rows, 1):
        r["priority"] = i
    return rows


def header_font():
    return Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")


def body_font():
    return Font(name=FONT_NAME, size=10, color=TEXT_DARK)


def muted_font():
    return Font(name=FONT_NAME, size=10, color=TEXT_MUTED, italic=True)


def title_font():
    return Font(name=FONT_NAME, size=18, bold=True, color=LAVENDER_DARK)


def subtitle_font():
    return Font(name=FONT_NAME, size=11, color=TEXT_MUTED, italic=True)


def header_fill():
    return PatternFill("solid", start_color=LAVENDER_DARK, end_color=LAVENDER_DARK)


def soft_fill():
    return PatternFill("solid", start_color=LAVENDER_LIGHTER, end_color=LAVENDER_LIGHTER)


def accent_fill():
    return PatternFill("solid", start_color=LAVENDER_LIGHT, end_color=LAVENDER_LIGHT)


def apply_header(ws, row, cols):
    for i, label in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=label)
        c.font = header_font()
        c.fill = header_fill()
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = thin_border


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_targets_sheet(wb, rows):
    ws = wb.create_sheet("Targets")

    ws["A1"] = "SoCal Midwife Outreach Targets"
    ws["A1"].font = title_font()
    ws.merge_cells("A1:T1")

    ws["A2"] = (
        "Combined pool: CALM directory + Supabase providers table, filtered to 8 SoCal counties. "
        "Tier 1 has email, Tier 2 has practice website, Tier 3 needs manual research."
    )
    ws["A2"].font = subtitle_font()
    ws.merge_cells("A2:T2")
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 18

    header_row = 4
    cols = [
        "Priority",
        "Tier",
        "Source",
        "First Name",
        "Last Name",
        "Practice / Organization",
        "License Type",
        "License #",
        "City",
        "Primary County",
        "All Counties Served",
        "Email",
        "Phone",
        "Website",
        "Medi-Cal?",
        "Last Contacted",
        "Status",
        "Practice Details",
        "Profile / Notes",
        "Free-form Notes",
    ]
    apply_header(ws, header_row, cols)
    ws.row_dimensions[header_row].height = 32

    tier_fills = {
        1: PatternFill("solid", start_color="E3F0DB", end_color="E3F0DB"),  # soft green
        2: PatternFill("solid", start_color="FCEFE0", end_color="FCEFE0"),  # soft peach
        3: PatternFill("solid", start_color="F3E9F7", end_color="F3E9F7"),  # soft lavender
    }

    start = header_row + 1
    for i, r in enumerate(rows):
        rr = start + i
        ws.cell(row=rr, column=1, value=r.get("priority"))
        ws.cell(row=rr, column=2, value=int(r.get("tier") or 3))
        ws.cell(row=rr, column=3, value=r.get("source") or "")
        ws.cell(row=rr, column=4, value=r.get("first") or "")
        ws.cell(row=rr, column=5, value=r.get("last") or "")
        ws.cell(row=rr, column=6, value=r.get("organization") or "")
        ws.cell(row=rr, column=7, value=r.get("license_type") or "")
        ws.cell(row=rr, column=8, value=r.get("license_no") or "")
        ws.cell(row=rr, column=9, value=r.get("city") or "")
        ws.cell(row=rr, column=10, value=r.get("primary_county") or "")
        ws.cell(row=rr, column=11, value=r.get("counties_all") or "")
        ws.cell(row=rr, column=12, value=r.get("email") or "")
        ws.cell(row=rr, column=13, value=r.get("phone") or "")
        ws.cell(row=rr, column=14, value=r.get("website") or "")
        ws.cell(row=rr, column=15, value=r.get("medi_cal") or "")
        ws.cell(row=rr, column=16, value=None)
        ws.cell(row=rr, column=17, value="Not contacted")
        ws.cell(row=rr, column=18, value=r.get("practice_details") or "")
        ws.cell(row=rr, column=19, value=r.get("profile_url") or "")
        ws.cell(row=rr, column=20, value=r.get("notes") or "")

        tier = int(r.get("tier") or 3)
        base_fill = tier_fills.get(tier)
        stripe_fill = soft_fill() if i % 2 == 0 and tier != 1 else None
        fill = base_fill if tier in (1, 2) else stripe_fill
        for col in range(1, 21):
            cell = ws.cell(row=rr, column=col)
            cell.font = body_font()
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = thin_border
            if fill:
                cell.fill = fill

        # Make the tier cell pop so scanning is easy
        tier_cell = ws.cell(row=rr, column=2)
        tier_cell.font = Font(name=FONT_NAME, size=10, bold=True, color=LAVENDER_DARK)
        tier_cell.alignment = Alignment(horizontal="center", vertical="center")

    set_col_widths(
        ws,
        [8, 6, 18, 12, 16, 28, 12, 10, 14, 16, 22, 30, 16, 24, 10, 14, 16, 40, 30, 30],
    )
    ws.freeze_panes = ws.cell(row=start, column=4)
    return ws, start, start + len(rows) - 1


def build_outreach_log_sheet(wb, target_count):
    ws = wb.create_sheet("Outreach Log")

    ws["A1"] = "Outreach Log"
    ws["A1"].font = title_font()
    ws.merge_cells("A1:K1")
    ws["A2"] = "One row per send. Log every touch so Follow-ups and Summary sheets stay honest."
    ws["A2"].font = subtitle_font()
    ws.merge_cells("A2:K2")
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 18

    header_row = 4
    cols = [
        "Date Sent",
        "Target Name",
        "Email",
        "Variant",
        "Subject",
        "Channel",
        "Opened?",
        "Replied?",
        "Outcome",
        "Next Action",
        "Notes",
    ]
    apply_header(ws, header_row, cols)
    ws.row_dimensions[header_row].height = 32

    # Seed 60 blank rows so the user can just start typing.
    seed_rows = max(60, target_count * 2)
    for i in range(seed_rows):
        rr = header_row + 1 + i
        fill = soft_fill() if i % 2 == 0 else None
        for col in range(1, 12):
            c = ws.cell(row=rr, column=col, value=None)
            c.font = body_font()
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c.border = thin_border
            if fill:
                c.fill = fill

    set_col_widths(ws, [14, 22, 30, 12, 38, 12, 10, 10, 20, 28, 40])
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    return ws, header_row + 1, header_row + seed_rows


def build_followups_sheet(wb, log_first_row, log_last_row):
    ws = wb.create_sheet("Follow-ups")

    ws["A1"] = "Follow-up Queue"
    ws["A1"].font = title_font()
    ws.merge_cells("A1:F1")
    ws["A2"] = "Formula-driven. Pulls the most recent entry per target from Outreach Log and shows days since last touch."
    ws["A2"].font = subtitle_font()
    ws.merge_cells("A2:F2")
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 18

    ws["A4"] = "Today"
    ws["A4"].font = Font(name=FONT_NAME, size=10, bold=True, color=TEXT_DARK)
    ws["B4"] = "=TODAY()"
    ws["B4"].font = body_font()
    ws["B4"].number_format = "yyyy-mm-dd"

    header_row = 6
    cols = [
        "Target Name",
        "Last Sent",
        "Days Since",
        "Last Variant",
        "Last Outcome",
        "Suggested Next",
    ]
    apply_header(ws, header_row, cols)
    ws.row_dimensions[header_row].height = 32

    # 40 formula rows. Each pulls from Outreach Log by name match using the
    # most recent row per target. We intentionally leave Target Name blank so
    # the user types a name and the rest auto-resolves. Formulas guard
    # against empty input with IFERROR wrappers so we never surface errors.
    log_range_name = f"'Outreach Log'!$B${log_first_row}:$B${log_last_row}"
    log_range_date = f"'Outreach Log'!$A${log_first_row}:$A${log_last_row}"
    log_range_variant = f"'Outreach Log'!$D${log_first_row}:$D${log_last_row}"
    log_range_outcome = f"'Outreach Log'!$I${log_first_row}:$I${log_last_row}"

    for i in range(40):
        rr = header_row + 1 + i
        ws.cell(row=rr, column=1, value=None)
        # Last Sent: MAX of dates where name matches
        ws.cell(
            row=rr,
            column=2,
            value=(
                f'=IFERROR(IF(A{rr}="","",'
                f"MAX(IF({log_range_name}=A{rr},{log_range_date}))),\"\")"
            ),
        )
        ws.cell(
            row=rr,
            column=3,
            value=(
                f'=IFERROR(IF(OR(A{rr}="",B{rr}=""),"",$B$4-B{rr}),"")'
            ),
        )
        # Last Variant: INDEX/MATCH the max date row
        ws.cell(
            row=rr,
            column=4,
            value=(
                f'=IFERROR(IF(A{rr}="","",'
                f"INDEX({log_range_variant},"
                f"MATCH(1,({log_range_name}=A{rr})*"
                f"({log_range_date}=B{rr}),0))),\"\")"
            ),
        )
        ws.cell(
            row=rr,
            column=5,
            value=(
                f'=IFERROR(IF(A{rr}="","",'
                f"INDEX({log_range_outcome},"
                f"MATCH(1,({log_range_name}=A{rr})*"
                f"({log_range_date}=B{rr}),0))),\"\")"
            ),
        )
        ws.cell(
            row=rr,
            column=6,
            value=(
                f'=IF(A{rr}="","",IF(C{rr}="","Log the first send",'
                f'IF(C{rr}<3,"Wait",IF(C{rr}<8,"Nudge variant B","Final touch or close"))))'
            ),
        )

        fill = soft_fill() if i % 2 == 0 else None
        for col in range(1, 7):
            c = ws.cell(row=rr, column=col)
            c.font = body_font()
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c.border = thin_border
            if fill:
                c.fill = fill
        ws.cell(row=rr, column=2).number_format = "yyyy-mm-dd"
        ws.cell(row=rr, column=3).number_format = "0"

    set_col_widths(ws, [24, 14, 14, 14, 22, 28])
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def build_variants_sheet(wb):
    ws = wb.create_sheet("Message Variants")

    ws["A1"] = "Cold Outreach: 4 Variants for A/B Testing"
    ws["A1"].font = title_font()
    ws.merge_cells("A1:D1")
    ws["A2"] = (
        "Merge fields: {{FirstName}}, {{Practice}}, {{CityOrCounty}}, {{LicenseType}}. "
        "Voice: Homebirth Provider Pitch. No em dashes anywhere."
    )
    ws["A2"].font = subtitle_font()
    ws.merge_cells("A2:D2")
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 22

    variants = [
        {
            "label": "Variant A: Founder Story",
            "hook": "Origin / why this exists",
            "subject": "A birth center closed on us at 24 weeks",
            "body": (
                "Hi {{FirstName}},\n\n"
                "Quick note from a fellow human, not a sales pitch.\n\n"
                "My co-founder Jo and I are building Homebirth.com. The short version of why: "
                "our friend Marie was 24 weeks pregnant in Knoxville when her birth center "
                "closed without warning. She had no midwife, no plan, and nowhere local to turn. "
                "A week later another friend, Niko, had to fly from Costa Rica to Miami for a "
                "homebirth because the options in her area felt opaque and scattered.\n\n"
                "We realized the problem was not a lack of great midwives. It was how impossible "
                "it is for parents to find the right one. So we bought the domain in 2019 and "
                "finally started building.\n\n"
                "We are launching in SoCal first, and I would love to learn what works and what "
                "is broken in the current discovery experience from someone actually in practice "
                "at {{Practice}}. Fifteen minutes on the phone, no ask, just listening.\n\n"
                "Would next Tuesday or Thursday work?\n\n"
                "Eddie Zaldivar\nFounder, Homebirth.com"
            ),
        },
        {
            "label": "Variant B: Product Difference",
            "hook": "Mirror intake / matching mechanics",
            "subject": "Matching parents to midwives the way it should work",
            "body": (
                "Hi {{FirstName}},\n\n"
                "I am building something I think you will have strong opinions on, and I want to "
                "hear them before we launch.\n\n"
                "Most midwife directories are alphabetical lists with a zip filter. Ours is not. "
                "Parents answer a 12 to 18 question intake about care style, birth setting, "
                "budget, values, and scope. Midwives answer the mirror version from the provider "
                "side. The platform runs hard filters first (radius, due date, payment) and then "
                "soft scores on fit, and every match comes with a short explanation like Within "
                "10 miles, open for June due dates, VBAC experienced, insurance compatible.\n\n"
                "The point is that you would only ever see parents whose actual needs align with "
                "your practice. No tire kickers. No leads outside your radius. No wasted consults.\n\n"
                "I am picking the first cohort of midwives to shape the mirror intake with us. "
                "{{Practice}} would be a great voice in that room. Free during beta, no strings.\n\n"
                "Open to a 15 minute call this week?\n\n"
                "Eddie Zaldivar\nFounder, Homebirth.com"
            ),
        },
        {
            "label": "Variant C: Ground Floor Partnership",
            "hook": "Equity tier, early builder seat",
            "subject": "Building Homebirth.com with a small group of founding midwives",
            "body": (
                "Hi {{FirstName}},\n\n"
                "I am the founder of Homebirth.com. We are putting together a small group of "
                "founding SoCal midwives to shape the platform before launch. Thought of {{Practice}} "
                "because you are exactly the kind of {{LicenseType}} practice we are building this for.\n\n"
                "Here is what that looks like. No cash. No subscription. Free listing plus a "
                "founding member seat that includes a small equity allocation from a pool we set "
                "aside specifically for midwives who help us get this right. The goal is simple: "
                "we only win if you win, and we want that aligned from day one.\n\n"
                "In exchange, we ask for real feedback, two or three calls across the build, and "
                "the chance to learn from how you actually run a practice. You keep every client "
                "you bring in through the platform at zero cost during beta, and your profile goes "
                "live the day we launch.\n\n"
                "Would you have 20 minutes in the next two weeks to talk it through?\n\n"
                "Eddie Zaldivar\nFounder, Homebirth.com"
            ),
        },
        {
            "label": "Variant D: Data-Driven Homework",
            "hook": "Medi-Cal gap insight, shows we did the work",
            "subject": "A gap I noticed in the SoCal midwife landscape",
            "body": (
                "Hi {{FirstName}},\n\n"
                "I have been doing homework on the out-of-hospital birth market in Southern "
                "California and noticed something that stopped me: of the 50 midwives in the "
                "California Association of Licensed Midwives directory, exactly one accepts "
                "Medi-Cal. One.\n\n"
                "That is a gap worth talking about, and also a signal that most of you have built "
                "private pay practices for a reason. Trust based. Relationship first. Not volume "
                "driven.\n\n"
                "That is the exact kind of practice we are building Homebirth.com for. Parents "
                "answer a structured intake about what they actually want in a midwife, we match "
                "on fit rather than proximity alone, and every lead arrives with context about why "
                "she chose you. The parents who land on {{Practice}} will already know why they "
                "are there.\n\n"
                "We are launching with a founding cohort in SoCal. I would love 15 minutes to hear "
                "how you think about lead quality and what would make this worth your time.\n\n"
                "Eddie Zaldivar\nFounder, Homebirth.com"
            ),
        },
    ]

    header_row = 4
    cols = ["Variant", "Hook", "Subject Line", "Body"]
    apply_header(ws, header_row, cols)
    ws.row_dimensions[header_row].height = 32

    for i, v in enumerate(variants):
        rr = header_row + 1 + i
        ws.cell(row=rr, column=1, value=v["label"])
        ws.cell(row=rr, column=2, value=v["hook"])
        ws.cell(row=rr, column=3, value=v["subject"])
        ws.cell(row=rr, column=4, value=v["body"])

        for col in range(1, 5):
            c = ws.cell(row=rr, column=col)
            c.font = body_font()
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c.border = thin_border
        ws.cell(row=rr, column=1).font = Font(name=FONT_NAME, size=10, bold=True, color=LAVENDER_DARK)
        ws.cell(row=rr, column=1).fill = soft_fill()
        ws.row_dimensions[rr].height = 320

    set_col_widths(ws, [22, 28, 44, 90])
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def build_summary_sheet(wb, targets_first_row, targets_last_row, log_first_row, log_last_row):
    ws = wb.create_sheet("Summary")

    ws["A1"] = "Outreach Dashboard"
    ws["A1"].font = title_font()
    ws.merge_cells("A1:E1")
    ws["A2"] = "All numbers are live formulas reading from Targets and Outreach Log."
    ws["A2"].font = subtitle_font()
    ws.merge_cells("A2:E2")
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 18

    label_font = Font(name=FONT_NAME, size=11, bold=True, color=TEXT_DARK)
    metric_font = Font(name=FONT_NAME, size=20, bold=True, color=LAVENDER_DARK)
    section_font = Font(name=FONT_NAME, size=13, bold=True, color=LAVENDER_DARK)

    def place(row, label, formula, fmt="0"):
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = label_font
        lc.alignment = Alignment(horizontal="left", vertical="center")
        vc = ws.cell(row=row, column=2, value=formula)
        vc.font = metric_font
        vc.number_format = fmt
        vc.alignment = Alignment(horizontal="left", vertical="center")

    # New Targets layout:
    #  A Priority  B Tier  C Source  D First  E Last  F Practice
    #  G License Type  H License #  I City  J Primary County
    #  K Counties  L Email  M Phone  N Website  O Medi-Cal
    #  P Last Contacted  Q Status  R Practice Details  S Profile  T Notes
    targets_priority_range = f"Targets!$A${targets_first_row}:$A${targets_last_row}"
    targets_tier_range = f"Targets!$B${targets_first_row}:$B${targets_last_row}"
    targets_source_range = f"Targets!$C${targets_first_row}:$C${targets_last_row}"
    targets_license_range = f"Targets!$G${targets_first_row}:$G${targets_last_row}"
    targets_county_range = f"Targets!$J${targets_first_row}:$J${targets_last_row}"
    targets_email_range = f"Targets!$L${targets_first_row}:$L${targets_last_row}"
    targets_website_range = f"Targets!$N${targets_first_row}:$N${targets_last_row}"
    targets_medi_range = f"Targets!$O${targets_first_row}:$O${targets_last_row}"

    ws.cell(row=4, column=1, value="TARGET ROSTER").font = section_font

    place(5, "Total targets", f"=COUNTA({targets_priority_range})")
    place(6, "With email on file", f'=COUNTIF({targets_email_range},"?*")')
    place(7, "With website only", f'=COUNTIF({targets_website_range},"?*")-COUNTIFS({targets_email_range},"?*",{targets_website_range},"?*")')
    place(8, "LM practices", f'=COUNTIF({targets_license_range},"LM*")')
    place(9, "Accept Medi-Cal", f'=COUNTIF({targets_medi_range},"Yes")')

    ws.cell(row=11, column=1, value="TIER BREAKDOWN").font = section_font
    place(12, "Tier 1 (email ready)", f"=COUNTIF({targets_tier_range},1)")
    place(13, "Tier 2 (website only)", f"=COUNTIF({targets_tier_range},2)")
    place(14, "Tier 3 (needs research)", f"=COUNTIF({targets_tier_range},3)")

    ws.cell(row=16, column=1, value="SOURCE BREAKDOWN").font = section_font
    place(17, "CALM directory", f'=COUNTIF({targets_source_range},"CALM")')
    place(18, "Supabase roster", f'=COUNTIF({targets_source_range},"Supabase")')
    place(19, "CALM + Supabase (confirmed)", f'=COUNTIF({targets_source_range},"CALM + Supabase")')

    ws.cell(row=21, column=1, value="TOP COUNTIES").font = section_font
    place(22, "Los Angeles", f'=COUNTIF({targets_county_range},"Los Angeles")')
    place(23, "San Diego", f'=COUNTIF({targets_county_range},"San Diego")')
    place(24, "Orange", f'=COUNTIF({targets_county_range},"Orange")')

    ws.cell(row=26, column=1, value="OUTREACH FUNNEL").font = section_font

    log_variant_range = f"'Outreach Log'!$D${log_first_row}:$D${log_last_row}"
    log_opened_range = f"'Outreach Log'!$G${log_first_row}:$G${log_last_row}"
    log_replied_range = f"'Outreach Log'!$H${log_first_row}:$H${log_last_row}"
    log_outcome_range = f"'Outreach Log'!$I${log_first_row}:$I${log_last_row}"

    place(27, "Total sends", f'=COUNTIF({log_variant_range},"?*")')
    place(28, "Opened", f'=COUNTIF({log_opened_range},"Yes")')
    place(29, "Replied", f'=COUNTIF({log_replied_range},"Yes")')
    place(30, "Meetings booked", f'=COUNTIF({log_outcome_range},"*Meeting*")')
    place(31, "Reply rate", f"=IFERROR(B29/B27,0)", "0.0%")

    ws.cell(row=33, column=1, value="VARIANT PERFORMANCE").font = section_font
    variant_header = ["Variant", "Sends", "Opens", "Replies", "Reply Rate"]
    for i, h in enumerate(variant_header, 1):
        c = ws.cell(row=34, column=i, value=h)
        c.font = header_font()
        c.fill = header_fill()
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = thin_border
    ws.row_dimensions[34].height = 24

    for i, letter in enumerate(["A", "B", "C", "D"]):
        rr = 35 + i
        ws.cell(row=rr, column=1, value=f"Variant {letter}").font = label_font
        ws.cell(
            row=rr,
            column=2,
            value=f'=COUNTIFS({log_variant_range},"{letter}")',
        ).number_format = "0"
        ws.cell(
            row=rr,
            column=3,
            value=(
                f'=COUNTIFS({log_variant_range},"{letter}",'
                f'{log_opened_range},"Yes")'
            ),
        ).number_format = "0"
        ws.cell(
            row=rr,
            column=4,
            value=(
                f'=COUNTIFS({log_variant_range},"{letter}",'
                f'{log_replied_range},"Yes")'
            ),
        ).number_format = "0"
        ws.cell(
            row=rr,
            column=5,
            value=f"=IFERROR(D{rr}/B{rr},0)",
        ).number_format = "0.0%"

        fill = soft_fill() if i % 2 == 0 else None
        for col in range(1, 6):
            cell = ws.cell(row=rr, column=col)
            cell.font = body_font() if col != 1 else label_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = thin_border
            if fill:
                cell.fill = fill

    ws.cell(row=40, column=1, value="Goal: reply rate above 15 percent across any single variant before scaling.").font = muted_font()
    ws.merge_cells("A40:E40")

    set_col_widths(ws, [30, 16, 14, 14, 16])
    for r in (5, 6, 7, 8, 9, 12, 13, 14, 17, 18, 19, 22, 23, 24, 27, 28, 29, 30, 31):
        ws.row_dimensions[r].height = 28


def main():
    if COMBINED_INPUT.exists():
        source_label = COMBINED_INPUT.name
    elif CALM_INPUT.exists():
        source_label = CALM_INPUT.name
    else:
        raise SystemExit(
            f"Missing input. Run scripts/dedupe-calm-vs-supabase.mjs "
            f"(preferred) or scripts/scrape-calm.mjs first. "
            f"Looked for: {COMBINED_INPUT.name}, {CALM_INPUT.name}."
        )

    rows = load_targets()
    print(f"Loaded {len(rows)} SoCal targets from {source_label}")

    # Tier + source breakdown so we know at a glance what we are working with
    tiers = {1: 0, 2: 0, 3: 0}
    sources = {}
    for r in rows:
        tiers[int(r.get("tier") or 3)] = tiers.get(int(r.get("tier") or 3), 0) + 1
        sources[r.get("source") or "Unknown"] = sources.get(r.get("source") or "Unknown", 0) + 1
    print(
        f"  Tier 1 (email): {tiers.get(1, 0)}  "
        f"Tier 2 (website): {tiers.get(2, 0)}  "
        f"Tier 3 (research): {tiers.get(3, 0)}"
    )
    for k, v in sorted(sources.items(), key=lambda x: -x[1]):
        print(f"  {k}: {v}")

    wb = Workbook()
    wb.remove(wb.active)

    targets_ws, tfirst, tlast = build_targets_sheet(wb, rows)
    log_ws, lfirst, llast = build_outreach_log_sheet(wb, len(rows))
    build_followups_sheet(wb, lfirst, llast)
    build_variants_sheet(wb)
    build_summary_sheet(wb, tfirst, tlast, lfirst, llast)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"Wrote {OUTPUT} ({OUTPUT.stat().st_size} bytes)")
    print(f"  Targets rows: {tfirst} to {tlast}")
    print(f"  Log rows:     {lfirst} to {llast}")


if __name__ == "__main__":
    main()
